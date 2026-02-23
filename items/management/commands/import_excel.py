"""
Command для импорта данных из Excel файла.
"""
import os
from datetime import datetime
from django.db import transaction
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
from items.models import Item, Expense


class Command(BaseCommand):
    help = 'Импорт данных из Excel файла'

    def add_arguments(self, parser):
        parser.add_argument(
            'file',
            type=str,
            help='Путь к Excel файлу'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Режим без записи в БД (только отчёт)'
        )
        parser.add_argument(
            '--no-clear',
            action='store_true',
            help='Не очищать базу данных перед импортом'
        )

    @transaction.atomic
    def handle(self, *args, **options):
        file_path = options['file']
        dry_run = options['dry_run']
        no_clear = options['no_clear']

        if not os.path.exists(file_path):
            raise CommandError(f'Файл {file_path} не найден')

        self.stdout.write(f'Загрузка файла: {file_path}')
        if dry_run:
            self.stdout.write(self.style.WARNING('РЕЖИМ DRY-RUN - данные не будут записаны'))

        # Очищаем базу данных
        if not no_clear:
            self.stdout.write('Очистка базы данных...')
            if not dry_run:
                Item.objects.all().delete()
                Expense.objects.all().delete()
            self.stdout.write(self.style.SUCCESS('База данных очищена'))

        # Загружаем Excel
        wb = load_workbook(file_path, data_only=True)

        total_items = 0
        total_expenses = 0
        date_formats = ['%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d-%m-%Y']

        for sheet_name in wb.sheetnames:
            self.stdout.write(f'Обработка листа: {sheet_name}')

            # Парсинг даты из названия листа
            purchase_date = self._parse_date_from_sheet_name(sheet_name, date_formats)
            if not purchase_date:
                self.stdout.write(self.style.WARNING(
                    f'  Не удалось распарсить дату из "{sheet_name}", пропускаем'
                ))
                continue

            sheet = wb[sheet_name]
            sheet_expense_amount = None

            for row_idx in range(2, sheet.max_row + 1):
                col_a = sheet.cell(row=row_idx, column=1).value

                # Проверка на "Накладные расходы"
                if col_a and str(col_a).strip().lower() == 'накладные расходы:':
                    expense_cell = sheet.cell(row=row_idx, column=5).value
                    if expense_cell:
                        expense_amount = self._parse_value(expense_cell)
                        if expense_amount is not None:
                            sheet_expense_amount = expense_amount
                            self.stdout.write(
                                self.style.SUCCESS(f'  Найдены накладные расходы: {expense_amount} ₽')
                            )
                    continue

                # Обработка предмета
                name = self._parse_name(sheet.cell(row=row_idx, column=2).value)
                if not name:
                    continue

                purchase_price = self._parse_value(sheet.cell(row=row_idx, column=3).value)
                if purchase_price is None:
                    self.stdout.write(self.style.WARNING(
                        f'  Неверная цена покупки в строке {row_idx}, пропускаем'
                    ))
                    continue

                sale_price = self._parse_value(sheet.cell(row=row_idx, column=4).value)
                sale_date = purchase_date if sale_price is not None else None

                if not dry_run:
                    Item.objects.create(
                        name=name,
                        purchase_price=purchase_price,
                        sale_price=sale_price,
                        purchase_date=purchase_date,
                        sale_date=sale_date
                    )
                total_items += 1

            # Добавляем накладные расходы
            if sheet_expense_amount is not None and not dry_run:
                Expense.objects.create(date=purchase_date, amount=sheet_expense_amount)
                total_expenses += 1

        self.stdout.write(self.style.SUCCESS(
            f'Импорт завершён! Предметов: {total_items}, расходов: {total_expenses}'
        ))
        wb.close()

    def _parse_date_from_sheet_name(self, sheet_name, formats):
        """Парсинг даты из названия листа."""
        clean_name = sheet_name.replace('(!)', '').replace('!', '').strip().split(' ')[0]

        for fmt in formats:
            try:
                return datetime.strptime(clean_name, fmt).date()
            except ValueError:
                continue
        return None

    def _parse_name(self, value):
        """Парсинг названия предмета."""
        if not value:
            return None
        name = str(value).strip()
        return name if name else None

    def _parse_value(self, value):
        """Парсинг числового значения."""
        if value is None or str(value).strip() == '':
            return None
        try:
            return int(float(str(value).replace(' ', '').replace('₽', '').strip()))
        except (ValueError, TypeError):
            return None
